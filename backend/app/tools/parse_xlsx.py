"""Parse .xlsx -> text via openpyxl."""
from __future__ import annotations

import os
from openpyxl import load_workbook


def _cell(v):
  if v is None:
    return ""
  if isinstance(v, float) and v.is_integer():
    return str(int(v))
  return str(v).strip()


def parse_xlsx(path: str) -> dict:
  """Extract all sheet text. Each sheet becomes a section."""
  if not os.path.isfile(path):
    raise FileNotFoundError(path)
  wb = load_workbook(path, data_only=True, read_only=True)
  parts = []
  for sheet in wb.worksheets:
    rows_out = []
    for row in sheet.iter_rows(values_only=True):
      cells = [_cell(v) for v in row]
      if any(cells):
        rows_out.append(" | ".join(cells))
    if rows_out:
      parts.append(f"[Sheet {sheet.title}]\n" + "\n".join(rows_out))
  wb.close()
  content = "\n\n".join(parts).strip()
  title = os.path.splitext(os.path.basename(path))[0] or "Untitled Spreadsheet"
  if not content:
    content = "(empty spreadsheet)"
  return {"title": title[:200], "content": content}
